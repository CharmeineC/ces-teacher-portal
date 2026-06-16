"""
app.py — Teacher Portal
Separate website for teachers to submit student enrollment data.
Hosted on Railway.app for free public access.

Admin credentials: admin / CES@2026
"""

from flask import (
    Flask, render_template, request, jsonify,
    redirect, url_for, session, Response
)
from werkzeug.utils import secure_filename
import os, csv, io, base64
from database import (
    setup_database, add_section, get_all_sections, get_section,
    update_section_adviser, get_all_grades, add_student, update_student,
    update_student_photo, update_id_status, get_students,
    get_student_by_id, bulk_import_students, get_stats,
    export_students_csv
)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "teacher_portal_ces_2026")

# Initialize database on startup
setup_database()

# ── Config ────────────────────────────────────────────────────────────────────
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "CES@2026")
SCHOOL_NAME    = os.environ.get("SCHOOL_NAME", "Communal Elementary School")

# For photo uploads — use local storage (Railway persistent disk)
# or set CLOUDINARY_URL env var to use Cloudinary
UPLOAD_FOLDER = os.environ.get("UPLOAD_FOLDER", "static/uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10MB


def allowed_image(filename):
    return "." in filename and \
           filename.rsplit(".", 1)[1].lower() in {"jpg", "jpeg", "png", "gif", "webp"}


def upload_photo(file, grade="", section_name="", student_name=""):
    """
    Save photo and return URL.
    Priority: Cloudinary → local storage fallback
    """
    if not file or not file.filename:
        return ""

    import time

    # Build clean filename
    if student_name:
        clean_name = student_name.upper().replace(" ", "_").replace(",", "")
    else:
        clean_name = str(int(time.time()))

    safe_name = secure_filename(file.filename)
    ext = safe_name.rsplit(".", 1)[-1].lower() if "." in safe_name else "jpg"
    if ext not in ["jpg","jpeg","png","gif","webp"]:
        ext = "jpg"
    filename = f"{clean_name}.{ext}"

    # Read file content once
    try:
        file_content = file.read()
    except Exception as e:
        print(f"Error reading file: {e}")
        return ""

    if not file_content:
        print("Upload: file content is empty!")
        return ""

    print(f"Upload: {filename} {len(file_content)} bytes grade={grade} section={section_name}")

    # Try Cloudinary first
    cloudinary_url = os.environ.get("CLOUDINARY_URL")
    if cloudinary_url:
        try:
            import cloudinary
            import cloudinary.uploader
            import io as _io

            # Folder organized by grade/section
            grade_c   = (grade or "general").replace(" ","").replace("/","-")
            section_c = (section_name or "general").replace(" ","").replace("/","-")
            folder    = f"ces_portal/{grade_c}-{section_c}"

            result = cloudinary.uploader.upload(
                _io.BytesIO(file_content),
                folder=folder,
                public_id=clean_name,
                overwrite=True,
                resource_type="image"
            )
            url = result.get("secure_url","")
            if url:
                print(f"Cloudinary upload success: {url}")
                return url
            else:
                print("Cloudinary returned no URL")
        except Exception as e:
            print(f"Cloudinary upload error: {e}")
            import traceback; traceback.print_exc()

    # Local storage fallback
    try:
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        filename_local = f"{int(time.time())}_{filename}"
        filepath = os.path.join(UPLOAD_FOLDER, filename_local)
        with open(filepath, "wb") as f_out:
            f_out.write(file_content)
        print(f"Saved locally: {filepath}")
        return f"/static/uploads/{filename_local}"
    except Exception as e:
        print(f"Local upload failed: {e}")
        try:
            import base64
            b64 = base64.b64encode(file_content).decode()
            return f"data:image/{ext};base64,{b64}"
        except Exception:
            return ""


def is_admin():
    return session.get("admin_logged_in") is True


# ── PUBLIC ROUTES ─────────────────────────────────────────────────────────────

@app.route("/")
def index():
    """Main teacher portal page."""
    from database import get_dashboard_stats, auto_create_sections_from_students, get_students_filtered
    auto_create_sections_from_students()
    grade        = request.args.get("grade", "")
    section      = request.args.get("section", "")
    search       = request.args.get("search", "")
    filter_type  = request.args.get("filter", "")

    # If a stat filter is active, show filtered students
    if filter_type:
        students = get_students_filtered(filter_type, grade or None, section or None)
    else:
        students = get_students(grade or None, section or None, search or None)

    sections = get_all_sections()
    grades   = get_all_grades()
    stats    = get_dashboard_stats()
    return render_template("index.html",
                           students=students,
                           sections=sections,
                           grades=grades,
                           stats=stats,
                           school_name=SCHOOL_NAME,
                           selected_grade=grade,
                           selected_section=section,
                           search=search,
                           filter_type=filter_type,
                           is_admin=is_admin())


@app.route("/add_section", methods=["POST"])
def route_add_section():
    grade        = request.form.get("grade", "").strip()
    section_name = request.form.get("section_name", "").strip()
    adviser_name = request.form.get("adviser_name", "").strip()

    # Handle signature upload
    sig_url = ""
    if "signature" in request.files:
        sig_file = request.files["signature"]
        if sig_file and sig_file.filename and allowed_image(sig_file.filename):
            sig_url = upload_photo(
                sig_file,
                grade=grade,
                section_name=section_name,
                student_name=f"SIGNATURE_{adviser_name.replace(' ','_')}"
            )

    section_pin = request.form.get("section_pin", "").strip()

    if grade and section_name:
        add_section(grade, section_name, adviser_name, sig_url, section_pin)

        # Handle optional CSV upload
        csv_msg = ""
        if "csv_file" in request.files:
            csv_file = request.files["csv_file"]
            fname = csv_file.filename if csv_file else ""
            print(f"CSV file received: '{fname}'")
            # Accept .csv and .CSV and any case
            if csv_file and fname and fname.lower().endswith(".csv"):
                try:
                    raw_bytes = csv_file.stream.read()
                    print(f"CSV bytes read: {len(raw_bytes)}")
                    # Try utf-8-sig first, then utf-8, then latin-1
                    for enc in ("utf-8-sig", "utf-8", "latin-1"):
                        try:
                            decoded = raw_bytes.decode(enc)
                            break
                        except Exception:
                            continue
                    stream = io.StringIO(decoded)
                    reader = csv.DictReader(stream)
                    headers = reader.fieldnames
                    print(f"CSV headers found: {headers}")
                    students = []
                    for row in reader:
                        # Normalize keys: lowercase + replace spaces with underscores
                        row_norm = {
                            k.lower().strip().replace(" ", "_"): v
                            for k, v in row.items() if k
                        }
                        raw_lrn = str(row_norm.get("lrn") or "").strip()
                        # Clean scientific notation
                        try:
                            if raw_lrn and ('e' in raw_lrn.lower() or '.' in raw_lrn):
                                raw_lrn = str(int(float(raw_lrn)))
                        except Exception:
                            pass
                        s = {
                            "lrn":           raw_lrn,
                            "first_name":    str(row_norm.get("first_name") or row_norm.get("firstname") or "").strip(),
                            "last_name":     str(row_norm.get("last_name") or row_norm.get("lastname") or "").strip(),
                            "middle_initial":str(row_norm.get("middle_initial") or row_norm.get("mi") or "").strip(),
                            "extension":     str(row_norm.get("extension") or row_norm.get("ext") or "").strip(),
                            "grade":         grade,
                            "section_name":  section_name,
                            "adviser_name":  adviser_name,
                            "emergency_contact_name":    str(row_norm.get("emergency_contact_name") or "").strip(),
                            "emergency_contact_address": str(row_norm.get("emergency_contact_address") or "").strip(),
                            "emergency_contact_number":  str(row_norm.get("emergency_contact_number") or "").strip(),
                        }
                        if s["first_name"] or s["last_name"]:
                            students.append(s)
                    print(f"Students parsed from CSV: {len(students)}")

                    # Detect duplicate LRN (scientific notation problem)
                    lrns = [s["lrn"] for s in students if s["lrn"]]
                    if lrns and len(set(lrns)) == 1 and len(lrns) > 1:
                        csv_msg = (
                            f" ⚠️ All {len(lrns)} students have the same LRN ({lrns[0]})! "
                            "Fix: In Google Sheets, click column A → Format → Number → Plain text "
                            "→ delete and re-enter LRN numbers → re-download CSV."
                        )
                    elif students:
                        csv_added, csv_updated, csv_errors = bulk_import_students(students)
                        csv_msg = f" {csv_added} students added, {csv_updated} updated."
                        if csv_errors:
                            csv_msg += f" Issues: {'; '.join(csv_errors[:2])}"
                    else:
                        csv_msg = " No valid student rows found in CSV."
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    csv_msg = f" CSV error: {str(e)}"
            else:
                print(f"CSV file skipped — filename: '{fname}' not ending in .csv")

    if grade and section_name:
        return redirect(url_for("index",
                                msg=f"Class {grade} - {section_name} saved!{csv_msg}",
                                success="1"))
    return redirect(url_for("index", msg="Section saved!", success="1"))


@app.route("/add_student", methods=["POST"])
def route_add_student():
    grade   = request.form.get("grade", "")
    section = request.form.get("section_name", "")

    data = {
        "lrn":                       request.form.get("lrn", "").strip(),
        "first_name":                request.form.get("first_name", "").strip(),
        "last_name":                 request.form.get("last_name", "").strip(),
        "middle_initial":            request.form.get("middle_initial", "").strip(),
        "extension":                 request.form.get("extension", "").strip(),
        "grade":                     grade,
        "section_name":              section,
        "adviser_name":              request.form.get("adviser_name", "").strip(),
        "emergency_contact_name":    request.form.get("emergency_contact_name", "").strip(),
        "emergency_contact_address": request.form.get("emergency_contact_address", "").strip(),
        "emergency_contact_number":  request.form.get("emergency_contact_number", "").strip(),
    }

    # Handle photo
    if "photo" in request.files:
        photo_file = request.files["photo"]
        if photo_file and photo_file.filename and allowed_image(photo_file.filename):
            student_name = f"{data.get('last_name','')}_{data.get('first_name','')}"
            data["photo_url"] = upload_photo(
                photo_file,
                grade=data.get("grade",""),
                section_name=data.get("section_name",""),
                student_name=student_name
            )

    success, msg = add_student(data)
    return redirect(url_for("index", grade=grade, section=section,
                            msg=msg, success="1" if success else "0"))


@app.route("/edit_student/<int:student_id>", methods=["POST"])
def route_edit_student(student_id):
    student = get_student_by_id(student_id)
    if not student:
        return redirect(url_for("index"))

    data = {
        "lrn":                       request.form.get("lrn", "").strip(),
        "first_name":                request.form.get("first_name", "").strip(),
        "last_name":                 request.form.get("last_name", "").strip(),
        "middle_initial":            request.form.get("middle_initial", "").strip(),
        "extension":                 request.form.get("extension", "").strip(),
        "grade":                     request.form.get("grade", "").strip(),
        "section_name":              request.form.get("section_name", "").strip(),
        "adviser_name":              request.form.get("adviser_name", "").strip(),
        "emergency_contact_name":    request.form.get("emergency_contact_name", "").strip(),
        "emergency_contact_address": request.form.get("emergency_contact_address", "").strip(),
        "emergency_contact_number":  request.form.get("emergency_contact_number", "").strip(),
    }

    success, msg = update_student(student_id, data)

    # Handle new photo
    if success and "photo" in request.files:
        photo_file = request.files["photo"]
        if photo_file and photo_file.filename and allowed_image(photo_file.filename):
            student_name = f"{data.get('last_name','')}_{data.get('first_name','')}"
            photo_url = upload_photo(
                photo_file,
                grade=data.get("grade",""),
                section_name=data.get("section_name",""),
                student_name=student_name
            )
            update_student_photo(student_id, photo_url)

    return redirect(url_for("index",
                            grade=data["grade"],
                            section=data["section_name"],
                            msg=msg,
                            success="1" if success else "0"))


@app.route("/upload_photo/<int:student_id>", methods=["POST"])
def route_upload_photo(student_id):
    """Upload or update a student's photo."""
    student = get_student_by_id(student_id)
    if not student:
        return jsonify({"success": False, "message": "Student not found"})

    if "photo" not in request.files:
        return jsonify({"success": False, "message": "No file provided"})

    photo_file = request.files["photo"]
    if not photo_file or not photo_file.filename:
        return jsonify({"success": False, "message": "No file selected"})

    if not allowed_image(photo_file.filename):
        return jsonify({"success": False, "message": "Invalid file type"})

    student = get_student_by_id(student_id)
    student_name = f"{student['last_name']}_{student['first_name']}" if student else ""
    photo_url = upload_photo(
        photo_file,
        grade=student["grade"] if student else "",
        section_name=student["section_name"] if student else "",
        student_name=student_name
    )
    update_student_photo(student_id, photo_url)
    return jsonify({"success": True, "photo_url": photo_url})


@app.route("/bulk_import", methods=["POST"])
def route_bulk_import():
    """Import students from CSV file."""
    grade   = request.form.get("grade", "").strip()
    section = request.form.get("section_name", "").strip()
    adviser = request.form.get("adviser_name", "").strip()

    if "file" not in request.files:
        return redirect(url_for("index", msg="No file selected", success="0"))

    file = request.files["file"]
    if not file or not file.filename:
        return redirect(url_for("index", msg="No file selected", success="0"))

    # Grade and section MUST come from the form — not from CSV
    if not grade or not section:
        return redirect(url_for("index",
            msg="Please select your grade and section before importing.",
            success="0"))

    # Get adviser info from the section record
    from database import get_section
    sec_info = get_section(grade, section)
    adviser = sec_info["adviser_name"] if sec_info else adviser

    def clean_lrn(val):
        """Convert LRN to clean string — handles scientific notation from Excel."""
        if not val:
            return ""
        val = str(val).strip()
        # Handle scientific notation e.g. 1.23457E+11
        try:
            if 'e' in val.lower() and ('+' in val or '-' in val.split('e')[-1]):
                val = str(int(float(val)))
            elif '.' in val and val.replace('.','').replace('-','').isdigit():
                val = str(int(float(val)))
        except Exception:
            pass
        return val.strip()

    try:
        stream  = io.StringIO(file.stream.read().decode("utf-8-sig"))
        reader  = csv.DictReader(stream)
        students = []
        for row in reader:
            raw_lrn = str(row.get("lrn") or row.get("LRN") or "").strip()
            s = {
                "lrn":           clean_lrn(raw_lrn),
                "first_name":    str(row.get("first_name") or row.get("First Name") or row.get("FIRST NAME") or "").strip(),
                "last_name":     str(row.get("last_name") or row.get("Last Name") or row.get("LAST NAME") or "").strip(),
                "middle_initial":str(row.get("middle_initial") or row.get("MI") or row.get("Middle Initial") or "").strip(),
                "extension":     str(row.get("extension") or row.get("Extension") or row.get("EXT") or "").strip(),
                # Always use grade/section from form — never from CSV
                "grade":         grade,
                "section_name":  section,
                "adviser_name":  adviser,
                "emergency_contact_name":    str(row_lower.get("emergency_contact_name") or row_lower.get("emergency_contact") or "").strip(),
                "emergency_contact_address": str(row_lower.get("emergency_contact_address") or row_lower.get("address") or "").strip(),
                "emergency_contact_number":  str(row_lower.get("emergency_contact_number") or row_lower.get("contact_number") or "").strip(),
            }
            if s["first_name"] or s["last_name"]:
                students.append(s)

        added, updated, errors = bulk_import_students(students)
        msg = f"Done! {added} new students added, {updated} existing students updated in {grade} - {section}."
        if errors:
            msg += f" Issues: {'; '.join(errors[:3])}"
        return redirect(url_for("section_detail",
                                grade=grade,
                                section_name=section,
                                msg=msg,
                                success="1"))
    except Exception as e:
        return redirect(url_for("index", msg=f"Import error: {str(e)}", success="0"))


@app.route("/download_csv")
def route_download_csv():
    """Download student data as CSV for import into attendance system."""
    grade   = request.args.get("grade", "") or None
    section = request.args.get("section", "") or None
    csv_data = export_students_csv(grade, section)
    filename = f"students_{grade or 'all'}_{section or 'all'}.csv"
    return Response(
        csv_data,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ── ADMIN ROUTES ──────────────────────────────────────────────────────────────

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session["admin_logged_in"] = True
            return redirect(url_for("index"))
        return render_template("admin_login.html",
                               error="Invalid username or password",
                               school_name=SCHOOL_NAME)
    return render_template("admin_login.html", school_name=SCHOOL_NAME)


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_logged_in", None)
    return redirect(url_for("index"))


@app.route("/admin/update_status/<int:student_id>", methods=["POST"])
def admin_update_status(student_id):
    if not is_admin():
        return jsonify({"success": False, "message": "Not authorized"})
    field = request.form.get("field")
    value = request.form.get("value") == "1"
    success = update_id_status(student_id, field, value)
    return jsonify({"success": success})


@app.route("/api/student/<int:student_id>")
def api_student(student_id):
    """Return student data as JSON for edit modal."""
    student = get_student_by_id(student_id)
    if not student:
        return jsonify({"error": "Not found"})
    return jsonify(dict(student))


@app.route("/api/stats")
def api_stats():
    return jsonify(get_stats())



@app.route("/section/<grade>/<section_name>")
def section_detail(grade, section_name):
    """Section detail page — shows adviser info, signature, and all students."""
    from database import get_section_detail, get_dashboard_stats, auto_create_sections_from_students, get_section
    auto_create_sections_from_students()
    section, students = get_section_detail(grade, section_name)
    # Get full section info including PIN for admin view
    if not section:
        section = get_section(grade, section_name)
    stats = get_dashboard_stats()
    sections = get_all_sections()
    grades = get_all_grades()
    return render_template("section_detail.html",
                           section=section,
                           students=students,
                           school_name=SCHOOL_NAME,
                           stats=stats,
                           sections=sections,
                           grades=grades,
                           selected_grade=grade,
                           selected_section=section_name,
                           is_admin=is_admin())


@app.route("/download_photo/<int:student_id>")
def download_photo(student_id):
    """Download a single student photo with proper filename."""
    try:
        student = get_student_by_id(student_id)
        if not student or not student["photo_url"]:
            return "No photo found for this student", 404

        # Build clean filename
        last  = (student["last_name"]    or "UNKNOWN").upper().replace(" ", "")
        first = (student["first_name"]   or "UNKNOWN").upper().replace(" ", "")
        grade = (student["grade"]        or "").upper().replace(" ", "").replace("-","")
        sec   = (student["section_name"] or "").upper().replace(" ", "")
        
        photo_url = student["photo_url"]
        
        # Get file extension
        url_path = photo_url.split("?")[0]  # remove query params
        ext = url_path.split(".")[-1].lower() if "." in url_path else "jpg"
        if ext not in ["jpg","jpeg","png","gif","webp"]:
            ext = "jpg"
        
        filename = f"{last}_{first}_{grade}_{sec}.{ext}"

        # Local file (saved in static/uploads)
        if photo_url.startswith("/static/"):
            # Try multiple base paths
            for base in [os.getcwd(), "/app", os.path.dirname(__file__)]:
                filepath = os.path.join(base, photo_url.lstrip("/"))
                if os.path.exists(filepath):
                    from flask import send_file
                    return send_file(
                        filepath,
                        as_attachment=True,
                        download_name=filename
                    )
            return "Photo file not found on server", 404

        # Remote URL (Cloudinary, etc.)
        import requests as rq
        resp = rq.get(photo_url, timeout=15)
        if resp.status_code == 200:
            from flask import Response
            ctype = resp.headers.get("content-type", "image/jpeg")
            return Response(
                resp.content,
                mimetype=ctype,
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                    "Content-Length": str(len(resp.content))
                }
            )
        return f"Could not fetch photo (status {resp.status_code})", 404

    except Exception as e:
        print(f"Download photo error: {e}")
        return f"Error: {str(e)}", 500


@app.route("/download_photos_zip/<grade>/<section_name>")
def download_photos_zip(grade, section_name):
    """Download all photos for a section as a ZIP file."""
    import zipfile, io, requests as req
    from database import get_section_detail

    _, students = get_section_detail(grade, section_name)

    zip_buffer = io.BytesIO()
    count = 0

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for student in students:
            if not student["photo_url"]:
                continue

            last  = (student["last_name"] or "").upper().replace(" ", "")
            first = (student["first_name"] or "").upper().replace(" ", "")
            g     = (student["grade"] or "").upper().replace(" ", "")
            sec   = (student["section_name"] or "").upper().replace(" ", "")
            ext   = student["photo_url"].split(".")[-1].lower()
            if ext not in ["jpg","jpeg","png","gif","webp"]:
                ext = "jpg"
            filename = f"{last}_{first}_{g}_{sec}.{ext}"

            photo_url = student["photo_url"]
            try:
                if photo_url.startswith("/static/"):
                    found = False
                    for base in [os.getcwd(), "/app", os.path.dirname(__file__)]:
                        filepath = os.path.join(base, photo_url.lstrip("/"))
                        if os.path.exists(filepath):
                            with open(filepath, "rb") as f:
                                zf.writestr(filename, f.read())
                                count += 1
                                found = True
                                break
                    if not found:
                        continue
                else:
                    resp = req.get(photo_url, timeout=10)
                    if resp.status_code == 200:
                        zf.writestr(filename, resp.content)
                        count += 1
            except Exception:
                continue

    if count == 0:
        return "No photos found for this section", 404

    zip_buffer.seek(0)
    zip_name = f"photos_{grade}_{section_name}.zip".replace(" ", "_")
    from flask import Response
    return Response(
        zip_buffer.getvalue(),
        mimetype="application/zip",
        headers={"Content-Disposition": f"attachment; filename={zip_name}"}
    )


@app.route("/download_template_xlsx")
def download_template_xlsx():
    """Generate Excel template with LRN column pre-formatted as text."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student List"

        # Headers
        headers = [
            'lrn', 'last_name', 'first_name', 'middle_initial',
            'extension', 'emergency_contact_name',
            'emergency_contact_address', 'emergency_contact_number'
        ]

        # Style header row
        header_fill = PatternFill(start_color="007a4d", end_color="007a4d", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Sample data rows
        samples = [
            ['123456789001', 'dela Cruz', 'Juan', 'A', '',
             'Rosa dela Cruz', '123 Rizal St Davao City', '09284553934'],
            ['123456789002', 'Reyes', 'Maria', 'B', '',
             'Pedro Reyes', '456 Mabini St Davao City', '09171234567'],
            ['123456789003', 'Santos', 'Jose', 'C', 'Jr.',
             'Ana Santos', '789 Bonifacio Ave', '09281234567'],
        ]
        for row_data in samples:
            ws.append(row_data)

        # Format entire LRN column (A) as TEXT to prevent scientific notation
        from openpyxl.styles import numbers
        for row in ws.iter_rows(min_row=1, max_row=500, min_col=1, max_col=1):
            for cell in row:
                cell.number_format = '@'  # @ means Text format in Excel

        # Set column widths
        col_widths = [18, 18, 18, 8, 8, 22, 30, 18]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Add instructions sheet
        ws2 = wb.create_sheet("Instructions")
        ws2['A1'] = "HOW TO USE THIS TEMPLATE"
        ws2['A1'].font = Font(bold=True, size=14, color="007a4d")
        instructions = [
            "",
            "1. Go to the 'Student List' sheet (tab at the bottom)",
            "2. The LRN column (column A) is already formatted as Text — do not change this",
            "3. Delete the sample data rows (rows 2, 3, 4) and enter your students",
            "4. Fill in one student per row",
            "5. Leave cells blank if information is not available",
            "",
            "TO UPLOAD TO GOOGLE SHEETS:",
            "1. Go to sheets.google.com",
            "2. Click File → Import → Upload this file",
            "3. Select 'Replace spreadsheet'",
            "4. Click Import data",
            "5. Your headers and data are already there!",
            "",
            "TO SAVE AS CSV FOR UPLOAD:",
            "In Excel: File → Save As → CSV (Comma delimited)",
            "In Google Sheets: File → Download → Comma Separated Values (.csv)",
            "",
            "IMPORTANT: Each student must have a unique LRN number!",
        ]
        for i, line in enumerate(instructions, 2):
            ws2[f'A{i}'] = line
        ws2.column_dimensions['A'].width = 60

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=CES_Student_Template.xlsx'}
        )
    except ImportError:
        return redirect(url_for('download_template_csv'))


@app.route("/download_template_csv")
def download_template_csv():
    """
    Download CSV template with LRN formatted as text.
    Uses tab-separated or quoted values to prevent Excel
    from converting LRN to scientific notation.
    """
    # Use quoted LRN values to keep them as text in Excel
    template = (
        "lrn,last_name,first_name,middle_initial,extension,"
        "emergency_contact_name,emergency_contact_address,emergency_contact_number\n"
        # LRN is quoted to force Excel to treat it as text
        "\"123456789001\",dela Cruz,Juan,A,,"
        "Rosa dela Cruz,123 Rizal St Davao City,09284553934\n"
        "\"123456789002\",Reyes,Maria,B,,"
        "Pedro Reyes,456 Mabini St Davao City,09171234567\n"
        "\"123456789003\",Santos,Jose,C,Jr.,"
        "Ana Santos,789 Bonifacio St Davao City,09281234567\n"
    )
    from flask import Response
    return Response(
        template,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=CES_Student_Template.csv"}
    )


@app.route("/api/test_drive")
def api_test_drive():
    """Test Google Drive connection."""
    from google_drive import test_connection
    success, message = test_connection()
    return jsonify({"success": success, "message": message})


@app.route("/delete_section/<grade>/<section_name>", methods=["POST"])
def route_delete_section(grade, section_name):
    """Delete a section — requires PIN or admin."""
    from database import delete_section, verify_section_pin
    pin_input = request.form.get("pin", "").strip()

    if not is_admin():
        valid, msg = verify_section_pin(grade, section_name, pin_input)
        if not valid:
            return redirect(url_for("index",
                                    msg=f"Cannot delete: {msg}",
                                    success="0"))

    delete_section(grade, section_name)
    return redirect(url_for("index",
                            msg=f"Section {grade} - {section_name} deleted.",
                            success="1"))


@app.route("/delete_student/<int:student_id>", methods=["POST"])
def route_delete_student(student_id):
    """Delete a student — requires section PIN or admin."""
    from database import delete_student, get_student_by_id, verify_section_pin
    student = get_student_by_id(student_id)
    if not student:
        return redirect(url_for("index"))

    grade   = student["grade"]
    section = student["section_name"]

    if not is_admin():
        pin_input = request.form.get("pin", "").strip()
        valid, msg = verify_section_pin(grade, section, pin_input)
        if not valid:
            return redirect(url_for("section_detail",
                                    grade=grade,
                                    section_name=section,
                                    msg=f"Cannot delete: {msg}",
                                    success="0"))

    delete_student(student_id)
    return redirect(url_for("section_detail",
                            grade=grade,
                            section_name=section,
                            msg="Student deleted.",
                            success="1"))


@app.route("/admin/restore")
def admin_restore_page():
    """Show restore page."""
    if not is_admin():
        return redirect(url_for("index", msg="Admin access required", success="0"))
    return render_template("restore.html", school_name=SCHOOL_NAME)


@app.route("/admin/bulk_restore", methods=["POST"])
def admin_bulk_restore():
    """Admin only — restore all students from full exported CSV."""
    if not is_admin():
        return redirect(url_for("index", msg="Admin access required", success="0"))

    if "file" not in request.files:
        return redirect(url_for("index", msg="No file selected", success="0"))

    f = request.files["file"]
    if not f or not f.filename:
        return redirect(url_for("index", msg="No file selected", success="0"))

    try:
        from database import get_connection
        raw  = f.stream.read()
        print(f"Restore: received {len(raw)} bytes")
        decoded = None
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                decoded = raw.decode(enc)
                print(f"Restore: decoded with {enc}")
                break
            except: continue

        if not decoded:
            return redirect(url_for("index", msg="Could not read file", success="0"))

        stream  = io.StringIO(decoded)
        reader  = csv.DictReader(stream)
        print(f"Restore: CSV headers = {reader.fieldnames}")
        students = []
        for row in reader:
            row_norm = {k.lower().strip().replace(" ","_"): v for k, v in row.items() if k}
            raw_lrn  = str(row_norm.get("lrn") or "").strip()
            try:
                if raw_lrn and ('e' in raw_lrn.lower() or '.' in raw_lrn):
                    raw_lrn = str(int(float(raw_lrn)))
            except Exception:
                pass

            # Handle id_printed/distributed from export
            def yn(val):
                return 1 if str(val).strip().lower() in ("yes","1","true") else 0

            s = {
                "lrn":                       raw_lrn,
                "first_name":                str(row_norm.get("first_name") or "").strip(),
                "last_name":                 str(row_norm.get("last_name") or "").strip(),
                "middle_initial":            str(row_norm.get("middle_initial") or "").strip(),
                "extension":                 str(row_norm.get("extension") or "").strip(),
                "grade":                     str(row_norm.get("grade") or "").strip(),
                "section_name":              str(row_norm.get("section_name") or "").strip(),
                "adviser_name":              str(row_norm.get("adviser_name") or "").strip(),
                "emergency_contact_name":    str(row_norm.get("emergency_contact_name") or "").strip(),
                "emergency_contact_address": str(row_norm.get("emergency_contact_address") or "").strip(),
                "emergency_contact_number":  str(row_norm.get("emergency_contact_number") or "").strip(),
                "photo_url":                 str(row_norm.get("photo_url") or "").strip(),
                "id_printed":                yn(row_norm.get("id_printed")),
                "id_distributed":            yn(row_norm.get("id_distributed")),
                "adviser_signature":         str(row_norm.get("adviser_signature") or "").strip(),
                "section_pin":               str(row_norm.get("section_pin") or "").strip(),
            }
            if s["first_name"] or s["last_name"]:
                students.append(s)

        if not students:
            return redirect(url_for("index", msg="No students found in CSV", success="0"))

        # Restore students with photo_url and id status preserved
        from database import get_connection
        from datetime import datetime
        conn = get_connection()
        added = 0; updated = 0
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for s in students:
            lrn     = s["lrn"]
            section = s["section_name"]
            existing = None
            if lrn:
                existing = conn.execute(
                    "SELECT id FROM students WHERE lrn=? AND LOWER(section_name)=LOWER(?)",
                    (lrn, section)
                ).fetchone()
            if existing:
                conn.execute("""
                    UPDATE students SET
                        first_name=?, last_name=?, middle_initial=?, extension=?,
                        grade=?, section_name=?, adviser_name=?,
                        emergency_contact_name=?, emergency_contact_address=?,
                        emergency_contact_number=?, photo_url=?,
                        id_printed=?, id_distributed=?, updated_at=?
                    WHERE lrn=? AND LOWER(section_name)=LOWER(?)
                """, (
                    s["first_name"], s["last_name"], s["middle_initial"], s["extension"],
                    s["grade"], s["section_name"], s["adviser_name"],
                    s["emergency_contact_name"], s["emergency_contact_address"],
                    s["emergency_contact_number"], s["photo_url"],
                    s["id_printed"], s["id_distributed"], now,
                    lrn, section
                ))
                updated += 1
            else:
                conn.execute("""
                    INSERT INTO students (
                        lrn, first_name, last_name, middle_initial, extension,
                        grade, section_name, adviser_name,
                        emergency_contact_name, emergency_contact_address,
                        emergency_contact_number, photo_url,
                        id_printed, id_distributed, created_at, updated_at
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    lrn, s["first_name"], s["last_name"], s["middle_initial"], s["extension"],
                    s["grade"], s["section_name"], s["adviser_name"],
                    s["emergency_contact_name"], s["emergency_contact_address"],
                    s["emergency_contact_number"], s["photo_url"],
                    s["id_printed"], s["id_distributed"], now, now
                ))
                added += 1
        conn.commit()
        conn.close()

        # Restore sections WITH signatures and PINs
        conn2 = get_connection()
        sections_restored = set()
        for s in students:
            key = (s["grade"], s["section_name"])
            if key not in sections_restored and s["grade"] and s["section_name"]:
                existing_sec = conn2.execute(
                    "SELECT id FROM sections WHERE LOWER(grade)=LOWER(?) AND LOWER(section_name)=LOWER(?)",
                    (s["grade"], s["section_name"])
                ).fetchone()
                sig = s.get("adviser_signature","") or ""
                pin = s.get("section_pin","") or ""
                if existing_sec:
                    # Update PIN and signature if they were in the backup
                    if sig or pin:
                        conn2.execute("""
                            UPDATE sections SET adviser_name=?,
                            adviser_signature=CASE WHEN ?!='' THEN ? ELSE adviser_signature END,
                            section_pin=CASE WHEN ?!='' THEN ? ELSE section_pin END
                            WHERE LOWER(grade)=LOWER(?) AND LOWER(section_name)=LOWER(?)
                        """, (s["adviser_name"], sig, sig, pin, pin, s["grade"], s["section_name"]))
                else:
                    conn2.execute("""
                        INSERT OR IGNORE INTO sections
                        (grade, section_name, adviser_name, adviser_signature, section_pin)
                        VALUES (?,?,?,?,?)
                    """, (s["grade"], s["section_name"], s["adviser_name"], sig, pin))
                sections_restored.add(key)
        conn2.commit()
        conn2.close()

        return redirect(url_for("index",
            msg=f"✅ Restore complete! {added} students added, {updated} updated. "
                f"{len(sections_restored)} sections restored with signatures and PINs.",
            success="1"))

    except Exception as e:
        import traceback; traceback.print_exc()
        return redirect(url_for("index", msg=f"Restore error: {str(e)}", success="0"))


# ── TEACHER PROFILE ROUTES ────────────────────────────────────────────────────

@app.route("/teachers")
def teachers_page():
    from database import get_all_teachers
    search = request.args.get("search","")
    teachers = get_all_teachers(search or None)
    return render_template("teachers.html",
                           teachers=teachers,
                           search=search,
                           school_name=SCHOOL_NAME,
                           is_admin=is_admin())


@app.route("/add_teacher", methods=["POST"])
def route_add_teacher():
    from database import add_teacher, update_teacher_photo
    data = {
        "last_name":      request.form.get("last_name","").strip(),
        "first_name":     request.form.get("first_name","").strip(),
        "middle_initial": request.form.get("middle_initial","").strip(),
        "employee_no":    request.form.get("employee_no","").strip(),
        "position":       request.form.get("position","").strip(),
        "birth_date":     request.form.get("birth_date","").strip(),
        "blood_type":     request.form.get("blood_type","").strip(),
        "address":        request.form.get("address","").strip(),
        "prc_number":     request.form.get("prc_number","").strip(),
        "tin":            request.form.get("tin","").strip(),
        "philhealth":     request.form.get("philhealth","").strip(),
        "gsis":           request.form.get("gsis","").strip(),
        "hdmf":           request.form.get("hdmf","").strip(),
        "advisory_class": request.form.get("advisory_class","").strip(),
        "ec_name":        request.form.get("ec_name","").strip(),
        "ec_number":      request.form.get("ec_number","").strip(),
    }
    ok, status, teacher_id = add_teacher(data)

    # Handle photo
    if ok and teacher_id and "photo" in request.files:
        photo_file = request.files["photo"]
        if photo_file and photo_file.filename and allowed_image(photo_file.filename):
            name = f"{data['last_name']}_{data['first_name']}"
            photo_url = upload_photo(photo_file, grade="teachers",
                                     section_name="profiles", student_name=name)
            if photo_url:
                update_teacher_photo(teacher_id, photo_url)

    msg = f"Teacher {data['last_name']}, {data['first_name']} {'added' if status=='added' else 'updated'}!"
    return redirect(url_for("teachers_page", msg=msg, success="1"))


@app.route("/edit_teacher/<int:teacher_id>", methods=["POST"])
def route_edit_teacher(teacher_id):
    from database import add_teacher, get_teacher_by_id
    teacher = get_teacher_by_id(teacher_id)
    if not teacher:
        return redirect(url_for("teachers_page", msg="Teacher not found", success="0"))
    data = {
        "last_name":      request.form.get("last_name","").strip(),
        "first_name":     request.form.get("first_name","").strip(),
        "middle_initial": request.form.get("middle_initial","").strip(),
        "employee_no":    request.form.get("employee_no","").strip(),
        "position":       request.form.get("position","").strip(),
        "birth_date":     request.form.get("birth_date","").strip(),
        "blood_type":     request.form.get("blood_type","").strip(),
        "address":        request.form.get("address","").strip(),
        "prc_number":     request.form.get("prc_number","").strip(),
        "tin":            request.form.get("tin","").strip(),
        "philhealth":     request.form.get("philhealth","").strip(),
        "gsis":           request.form.get("gsis","").strip(),
        "hdmf":           request.form.get("hdmf","").strip(),
        "advisory_class": request.form.get("advisory_class","").strip(),
        "ec_name":        request.form.get("ec_name","").strip(),
        "ec_number":      request.form.get("ec_number","").strip(),
    }
    # Force update by using the existing employee_no if not provided
    if not data["employee_no"] and teacher["employee_no"]:
        data["employee_no"] = teacher["employee_no"]

    add_teacher(data)
    return redirect(url_for("teachers_page",
                            msg=f"Teacher {data['last_name']}, {data['first_name']} updated!",
                            success="1"))


@app.route("/upload_teacher_photo/<int:teacher_id>", methods=["POST"])
def route_upload_teacher_photo(teacher_id):
    from database import get_teacher_by_id, update_teacher_photo
    teacher = get_teacher_by_id(teacher_id)
    if not teacher:
        return jsonify({"success": False, "message": "Not found"})
    if "photo" not in request.files:
        return jsonify({"success": False, "message": "No file"})
    photo_file = request.files["photo"]
    if not photo_file or not photo_file.filename:
        return jsonify({"success": False, "message": "No file selected"})
    name = f"{teacher['last_name']}_{teacher['first_name']}"
    photo_url = upload_photo(photo_file, grade="teachers",
                             section_name="profiles", student_name=name)
    if photo_url:
        update_teacher_photo(teacher_id, photo_url)
        return jsonify({"success": True, "photo_url": photo_url})
    return jsonify({"success": False, "message": "Upload failed"})


@app.route("/delete_teacher/<int:teacher_id>", methods=["POST"])
def route_delete_teacher(teacher_id):
    if not is_admin():
        return redirect(url_for("teachers_page", msg="Admin only", success="0"))
    from database import delete_teacher, get_teacher_by_id
    t = get_teacher_by_id(teacher_id)
    if t:
        delete_teacher(teacher_id)
        return redirect(url_for("teachers_page",
                                msg=f"Teacher {t['last_name']}, {t['first_name']} deleted.",
                                success="1"))
    return redirect(url_for("teachers_page"))


@app.route("/import_teachers", methods=["POST"])
def route_import_teachers():
    if not is_admin():
        return redirect(url_for("teachers_page", msg="Admin only", success="0"))
    if "file" not in request.files:
        return redirect(url_for("teachers_page", msg="No file", success="0"))
    f = request.files["file"]
    try:
        raw = f.stream.read()
        for enc in ("utf-8-sig","utf-8","latin-1"):
            try: decoded = raw.decode(enc); break
            except: continue
        stream = io.StringIO(decoded)
        reader = csv.DictReader(stream)
        rows = []
        for row in reader:
            rn = {k.lower().strip().replace(" ","_"): v for k,v in row.items() if k}
            rows.append({
                "last_name":      str(rn.get("last_name","")).strip(),
                "first_name":     str(rn.get("first_name","")).strip(),
                "middle_initial": str(rn.get("middle_initial","")).strip(),
                "employee_no":    str(rn.get("employee_no","")).strip(),
                "position":       str(rn.get("position","")).strip(),
                "birth_date":     str(rn.get("birth_date","")).strip(),
                "blood_type":     str(rn.get("blood_type","")).strip(),
                "address":        str(rn.get("address","")).strip(),
                "prc_number":     str(rn.get("prc_number","")).strip(),
                "tin":            str(rn.get("tin","")).strip(),
                "philhealth":     str(rn.get("philhealth","")).strip(),
                "gsis":           str(rn.get("gsis","")).strip(),
                "hdmf":           str(rn.get("hdmf","")).strip(),
                "advisory_class": str(rn.get("advisory_class","")).strip(),
                "ec_name":        str(rn.get("ec_name","")).strip(),
                "ec_number":      str(rn.get("ec_number","")).strip(),
            })
        from database import bulk_import_teachers
        added, updated, errors = bulk_import_teachers(rows)
        msg = f"Done! {added} teachers added, {updated} updated."
        if errors: msg += f" Issues: {'; '.join(errors[:3])}"
        return redirect(url_for("teachers_page", msg=msg, success="1"))
    except Exception as e:
        return redirect(url_for("teachers_page", msg=f"Import error: {e}", success="0"))


@app.route("/download_teachers_csv")
def route_download_teachers_csv():
    from database import export_teachers_csv
    csv_data = export_teachers_csv()
    return Response(csv_data, mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=teachers.csv"})


@app.route("/api/teacher/<int:teacher_id>")
def api_teacher(teacher_id):
    from database import get_teacher_by_id
    t = get_teacher_by_id(teacher_id)
    if not t: return jsonify({"error":"Not found"})
    return jsonify(dict(t))

# ── ENTRY POINT ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    setup_database()
    port = int(os.environ.get("PORT", 8080))
    print(f"\n{'='*50}")
    print(f"  Teacher Portal — {SCHOOL_NAME}")
    print(f"{'='*50}")
    print(f"  URL: http://localhost:{port}")
    print(f"  Admin: http://localhost:{port}/admin/login")
    print(f"{'='*50}\n")
    app.run(host="0.0.0.0", port=port, debug=False)
